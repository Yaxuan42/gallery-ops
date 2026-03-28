"""
inventory.json → 现货清单.xlsx 同步脚本

Agent 更新 inventory.json 后运行此脚本生成 Excel 视图。
用法：python3 sync-excel.py
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(SCRIPT_DIR, "data", "inventory.json")
XLSX_PATH = os.path.join(SCRIPT_DIR, "data", "现货清单.xlsx")

# Column definitions: (group, field_key, header_label, width)
# field_key uses dot notation for nested fields
COLUMNS = [
    ("基本信息", "id",              "产品编号", 14),
    ("基本信息", "name",            "名称",     28),
    ("基本信息", "nickname",        "俗称",     12),
    ("基本信息", "model",           "型号",     16),
    ("基本信息", "category",        "产品分类", 14),
    ("基本信息", "designer",        "设计师",   20),
    ("基本信息", "origin",          "产地",     20),
    ("年代",     "design_year",     "设计年代", 12),
    ("年代",     "production_year", "生产年份", 12),
    ("规格",     "materials",       "材质",     16),
    ("规格",     "dimensions",      "尺寸",     16),
    ("规格",     "condition",       "品相等级", 10),
    ("价格",     "pricing.retail",      "定价",     12),
    ("价格",     "pricing.acquisition", "买手价",   12),
    ("价格",     "pricing.reference",   "参考公价", 12),
    ("溯源",     "provenance.institution",    "认证机构", 16),
    ("溯源",     "provenance.region",         "来源区域", 14),
    ("溯源",     "provenance.previous_owner", "前使用者", 14),
    ("溯源",     "provenance.channel",        "购入渠道", 14),
    ("内容",     "content.design_language", "设计语言", 30),
    ("内容",     "content.details",         "产品细节", 30),
    ("内容",     "content.story",           "背景故事", 30),
    ("管理",     "status", "状态", 10),
    ("管理",     "photos", "照片", 14),
    ("管理",     "notes",  "备注", 24),
]

GROUP_COLORS = {
    "基本信息": "4A4A48",
    "年代":     "6B6B68",
    "规格":     "4A4A48",
    "价格":     "6B6B68",
    "溯源":     "4A4A48",
    "内容":     "6B6B68",
    "管理":     "4A4A48",
}


def get_nested(obj, key):
    """Get value from nested dict using dot notation."""
    parts = key.split(".")
    val = obj
    for p in parts:
        if isinstance(val, dict):
            val = val.get(p)
        else:
            return None
    return val


def format_value(val):
    """Convert a value to display string."""
    if val is None:
        return ""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


def sync():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    items = data.get("items", [])

    # Collect extra keys across all items
    extra_keys = []
    for item in items:
        for k in item.get("extra", {}):
            if k not in extra_keys:
                extra_keys.append(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "现货清单"

    # Styles
    group_font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    header_font = Font(name="Inter", size=9, bold=True, color="1C1B19")
    header_fill = PatternFill("solid", fgColor="E8E8E4")
    data_font = Font(name="Inter", size=9, color="1C1B19")
    thin_border = Border(
        left=Side(style="thin", color="D0D0CC"),
        right=Side(style="thin", color="D0D0CC"),
        top=Side(style="thin", color="D0D0CC"),
        bottom=Side(style="thin", color="D0D0CC"),
    )

    all_columns = list(COLUMNS)
    for ek in extra_keys:
        all_columns.append(("扩展", f"extra.{ek}", ek, 16))

    # Row 1: Group headers
    col = 1
    groups_seen = {}
    for group, *_ in all_columns:
        if group not in groups_seen:
            groups_seen[group] = {"start": col, "end": col}
        else:
            groups_seen[group]["end"] = col
        col += 1

    for group, span in groups_seen.items():
        if span["start"] != span["end"]:
            ws.merge_cells(
                start_row=1, start_column=span["start"],
                end_row=1, end_column=span["end"]
            )
        color = GROUP_COLORS.get(group, "4A4A48")
        for c in range(span["start"], span["end"] + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = thin_border
        cell = ws.cell(row=1, column=span["start"], value=group)
        cell.font = group_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Column headers
    for i, (_, _, label, width) in enumerate(all_columns, 1):
        cell = ws.cell(row=2, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Data rows
    for row_idx, item in enumerate(items, 3):
        for col_idx, (_, key, _, _) in enumerate(all_columns, 1):
            val = get_nested(item, key)
            cell = ws.cell(row=row_idx, column=col_idx, value=format_value(val))
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "C3"

    wb.save(XLSX_PATH)
    print(f"✓ 同步完成：{len(items)} 件产品 → {XLSX_PATH}")
    print(f"  更新时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")


if __name__ == "__main__":
    sync()
